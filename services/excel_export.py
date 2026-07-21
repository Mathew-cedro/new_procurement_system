import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as database_config

# Category color fills
cat_colors = {
    # Project Status
    "initiated": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # Light Blue
    "bidding": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),    # Light Yellow
    "planning & bidding": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light Yellow
    "contract awarded": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # Light Gray
    "delivering": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Light Peach/Orange
    "deliveries & payments": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Light Peach
    "under warranty": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Light Green
    
    # Mode of Procurement
    "shopping": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "public bidding": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "direct contracting": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # Light Salmon
    "negotiated procurement": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),  # Light Purple
    "small value procurement": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    
    # Payment Type
    "progress payment": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "advance payment (15%)": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "final payment": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "one-time full payment": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    
    # Warranty Status
    "ongoing": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "released": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "pending claims": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "coa disallowance hold": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    
    # Delivery Status (compatibility)
    "pending": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "delivered": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "delayed": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "cancelled": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

def create_formatted_sheet(sheet, headers, rows):
    # Styling configurations
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write header columns
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Freeze panes below headers
    sheet.freeze_panes = "A2"
    
    # Data row style
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    # Write row records
    for row_idx, row_val in enumerate(rows, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(row_val, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            
            # Align and format based on column headers
            h_name = headers[col_idx - 1]
            
            # Category color-coding
            if h_name in ["Project Status", "Mode of Procurement", "Payment Type", "Warranty Status", "Delivery Status"]:
                if val:
                    val_str = str(val).strip().lower()
                    if val_str in cat_colors:
                        cell.fill = cat_colors[val_str]
            
            if h_name.endswith("Link"):
                cell.alignment = center_align
                if val:
                    cell.value = "Open PDF"
                    if not (val.startswith("http://") or val.startswith("https://") or val.startswith("\\\\") or (len(val) > 1 and val[1] == ":")):
                        import os
                        base_dir = os.path.dirname(os.path.abspath(__file__))
                        abs_val = os.path.normpath(os.path.join(base_dir, val))
                    else:
                        abs_val = val
                    cell.hyperlink = abs_val
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                else:
                    cell.value = "—"
            elif "(₱)" in h_name or "Allocation" in h_name or "Amount" in h_name or "Gross" in h_name or "Net" in h_name or "Savings" in h_name or "Price" in h_name:
                cell.alignment = right_align
                if val is not None:
                    try:
                        cell.value = float(val)
                    except (ValueError, TypeError):
                        pass
                    cell.number_format = '"₱"#,##0.00'
            elif "Date" in h_name or "ID" in h_name or "Serial" in h_name or "Status" in h_name or "No" in h_name:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Row Heights
    sheet.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        sheet.row_dimensions[r].height = 20
        
    # Auto-fit column widths
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if cell.number_format == '"₱"#,##0.00' and isinstance(cell.value, (int, float)):
                val = f"₱{cell.value:,.2f}"
            max_len = max(max_len, len(val))
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def get_base_query():
    return """
        SELECT 
            p.project_id, -- index 0
            p.project_id, p.project_name, p.bureau_division, p.focal_person, p.mode_of_procurement, p.approved_budget_abc, p.status,
            'MOOE' AS budget_type, p.approved_budget_abc AS app_amount, c.contract_amount AS ors_amount, 
            CASE WHEN c.contract_id IS NOT NULL THEN 'ORS-' || c.contract_id ELSE NULL END AS ors_serial_no,
            s.supplier_name,
            c.contract_id AS po_jo_contract_no, c.contract_amount,
            CASE WHEN c.contract_id IS NOT NULL THEN COALESCE(p.approved_budget_abc - c.contract_amount, 0.0) ELSE 0.0 END AS savings,
            d.milestone_description, d.delivery_status, d.actual_delivery_date,
            d.payment_type, d.payment_gross_amount, d.payment_net_amount, d.dv_rci_serial_no, d.check_date,
            w.warranty_status,
            COALESCE(c.ntp_date, p.date_received_bacsec, '2026-01-01') AS project_date,
            p.saro_pdf, p.ppmp_pdf, c.request_order_pdf, p.market_scoping_pdf, p.tech_specs_pdf,
            c.signed_contract_pdf, d.po_pdf, p.abstract_quotations_pdf, w.warranty_certificate_pdf
        FROM projects p
        LEFT JOIN contracts c ON p.project_id = c.project_id
        LEFT JOIN suppliers s ON c.supplier_id = s.supplier_id
        LEFT JOIN deliveries_and_payments d ON c.contract_id = d.contract_id
        LEFT JOIN warranties w ON c.contract_id = w.contract_id
    """

def export_master_data(output_path, filters=None):
    """
    Queries all tables from procurement.db, joins them into a flat format,
    applies optional filters (status, division, date range), and exports
    them to a styled native Excel worksheet (.xlsx).
    """
    conn = database_config.get_db_connection()
    cur = conn.cursor()
    
    query = get_base_query() + " ORDER BY p.project_id"
    
    try:
        cur.execute(query)
        rows = cur.fetchall()
        
        # Apply filters programmatically
        filtered_rows = []
        from datetime import datetime, date
        current_date = date.today()
        
        status_filter = filters.get("status", "All Statuses") if filters else "All Statuses"
        division_filter = filters.get("division", "All Divisions") if filters else "All Divisions"
        date_filter = filters.get("date", "All Dates") if filters else "All Dates"
        
        # Sort rows by date if chosen
        if date_filter == "Newest First" or date_filter == "Newest":
            rows = sorted(rows, key=lambda x: x[25] or "", reverse=True)
        elif date_filter == "Oldest First" or date_filter == "Oldest":
            rows = sorted(rows, key=lambda x: x[25] or "", reverse=False)
        
        for row in rows:
            row_status = row[7]
            row_division = row[3]
            row_date_str = row[25]
            
            # Status Filter
            if status_filter != "All Statuses" and row_status != status_filter:
                continue
                
            # Division Filter
            if division_filter != "All Divisions" and (row_division is None or row_division.strip() != division_filter.strip()):
                continue
                
            # Date Presets Filter
            if date_filter not in ("All Dates", "Newest First", "Oldest First") and row_date_str:
                try:
                    proj_date = datetime.strptime(row_date_str, "%Y-%m-%d").date()
                    days_diff = (current_date - proj_date).days
                    if date_filter == "Within 24 hours only" and not (0 <= days_diff <= 1):
                        continue
                    elif date_filter == "Within a week only" and not (0 <= days_diff <= 7):
                        continue
                    elif date_filter == "Within a month only" and not (0 <= days_diff <= 30):
                        continue
                except Exception:
                    pass
            
            # Form base data array
            base_data = list(row[1:25]) # project_id to warranty_status (including savings at index 15)
            base_data.extend(list(row[26:])) # saro_pdf to warranty_certificate_pdf
            filtered_rows.append(base_data)
                
        headers = [
            "Project ID", "Project Name", "Bureau/Division", "Focal Person", "Mode of Procurement", "ABC Budget (₱)", "Project Status",
            "Budget Type", "App Allocation (₱)", "ORS Amount (₱)", "ORS Serial No",
            "Supplier Name", "PO/JO Contract No", "Contract Price (₱)", "Savings (₱)",
            "Milestone Deliverable", "Delivery Status", "Actual Delivery Date",
            "Payment Type", "Payment Gross (₱)", "Payment Net (₱)", "DV / RCI No", "Check Date",
            "Warranty Status",
            "SARO Link", "PPMP Link", "RQ Link", "MS Link", "TS Link", "PO Phase 2 Link", "RFQ Phase 3 Link", "Abstract Link", "PO Phase 6 Link"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Procurement Registry"
        ws.views.sheetView[0].showGridLines = True
        
        create_formatted_sheet(ws, headers, filtered_rows)
        wb.save(output_path)
        
        return True, len(filtered_rows)
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()

def export_project_timeline(project_id, output_path):
    """
    Queries all timeline data related to a specific project and exports it to a styled native Excel worksheet (.xlsx).
    """
    conn = database_config.get_db_connection()
    cur = conn.cursor()
    
    query = get_base_query() + " WHERE p.project_id = ? ORDER BY p.project_id"
    
    try:
        cur.execute(query, (project_id,))
        rows = cur.fetchall()
        
        filtered_rows = []
        for row in rows:
            base_data = list(row[1:25])
            base_data.extend(list(row[26:]))
            filtered_rows.append(base_data)
            
        headers = [
            "Project ID", "Project Name", "Bureau/Division", "Focal Person", "Mode of Procurement", "ABC Budget (₱)", "Project Status",
            "Budget Type", "App Allocation (₱)", "ORS Amount (₱)", "ORS Serial No",
            "Supplier Name", "PO/JO Contract No", "Contract Price (₱)", "Savings (₱)",
            "Milestone Deliverable", "Delivery Status", "Actual Delivery Date",
            "Payment Type", "Payment Gross (₱)", "Payment Net (₱)", "DV / RCI No", "Check Date",
            "Warranty Status",
            "SARO Link", "PPMP Link", "RQ Link", "MS Link", "TS Link", "PO Phase 2 Link", "RFQ Phase 3 Link", "Abstract Link", "PO Phase 6 Link"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Timeline"
        ws.views.sheetView[0].showGridLines = True
        
        create_formatted_sheet(ws, headers, filtered_rows)
        wb.save(output_path)
        
        return True, len(filtered_rows)
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()
