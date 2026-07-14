import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Category color fills
cat_colors = {
    # Project Status
    "initiated": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # Light Blue
    "bidding": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),    # Light Yellow
    "contract awarded": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # Light Gray
    "delivering": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Light Peach/Orange
    "under warranty": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Light Green
    
    # Mode of Procurement
    "shopping": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "public bidding": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "direct contracting": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # Light Salmon
    "negotiated procurement": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),  # Light Purple
    "small value procurement": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    
    # Payment Type
    "progress payment": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "advance payment (15%)": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "final payment": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "one-time full payment": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    
    # Warranty Status
    "ongoing": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "released": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "pending claims": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "coa disallowance hold": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    
    # Delivery Status (compatibility)
    "pending": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "delivered": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "delayed": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "cancelled": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

def create_formatted_sheet(sheet, headers, rows):
    # Styling configurations
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write header columns
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Freeze panes below headers
    sheet.freeze_panes = "A2"
    
    # Data row style
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    # Write row records
    for row_idx, row_val in enumerate(rows, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(row_val, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            
            # Align and format based on column headers
            h_name = headers[col_idx - 1]
            
            # Category color-coding
            if h_name in ["Project Status", "Mode of Procurement", "Payment Type", "Warranty Status", "Delivery Status"]:
                if val:
                    val_str = str(val).strip().lower()
                    if val_str in cat_colors:
                        cell.fill = cat_colors[val_str]
            
            if h_name.endswith("Link"):
                cell.alignment = center_align
                if val:
                    cell.value = "Open PDF"
                    if not (val.startswith("http://") or val.startswith("https://") or val.startswith("\\\\") or (len(val) > 1 and val[1] == ":")):
                        import os
                        base_dir = os.path.dirname(os.path.abspath(__file__))
                        abs_val = os.path.normpath(os.path.join(base_dir, val))
                    else:
                        abs_val = val
                    cell.hyperlink = abs_val
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                else:
                    cell.value = "—"
            elif "(₱)" in h_name or "Allocation" in h_name or "Amount" in h_name or "Gross" in h_name or "Net" in h_name:
                cell.alignment = right_align
                if val is not None:
                    try:
                        cell.value = float(val)
                    except (ValueError, TypeError):
                        pass
                    cell.number_format = '"₱"#,##0.00'
            elif "Date" in h_name or "ID" in h_name or "Serial" in h_name or "Status" in h_name or "No" in h_name:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Row Heights
    sheet.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        sheet.row_dimensions[r].height = 20
        
    # Auto-fit column widths
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if cell.number_format == '"₱"#,##0.00' and isinstance(cell.value, (int, float)):
                val = f"₱{cell.value:,.2f}"
            max_len = max(max_len, len(val))
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def export_master_data(output_path, filters=None):
    """
    Queries all tables from procurement.db, joins them into a flat format,
    applies optional filters (status, division, date range), and exports
    them to a styled native Excel worksheet (.xlsx).
    """
    db_file = Path(__file__).parent / "procurement.db"
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    
    # Query project documents map
    try:
        cur.execute("SELECT project_id, document_type, file_reference FROM project_documents")
        project_docs_map = {}
        for pid, dtype, ref in cur.fetchall():
            if pid not in project_docs_map:
                project_docs_map[pid] = {}
            project_docs_map[pid][dtype] = ref
    except Exception:
        project_docs_map = {}
        
    query = """
        SELECT 
            p.id, -- index 0
            p.proj_id_no, p.project_name, p.bureau_division_name, p.focal_person, p.mode_of_procurement, p.abc_amount, p.status,
            pb.budget_type, pb.app_amount, pb.ors_amount, pb.ors_serial_no,
            s.supplier_name,
            c.po_jo_contract_no, c.contract_amount,
            d.milestone_deliverable, d.status_of_delivery, d.actual_delivery_date,
            pm.types_of_payment, pm.payment_amount_gross, pm.payment_amount_net, pm.rci_dv_acic_no, pm.check_date_lddap_ada,
            w.retention_period_warranty_status,
            COALESCE(c.contract_date, 
                     (SELECT MIN(d.date_prepared) FROM project_documents d WHERE d.project_id = p.id),
                     (SELECT MIN(b.date_received_bacsec) FROM bids b WHERE b.project_id = p.id),
                     '2026-01-01') AS project_date
        FROM projects p
        LEFT JOIN project_budgets pb ON p.id = pb.project_id
        LEFT JOIN contracts c ON p.id = c.project_id
        LEFT JOIN suppliers s ON c.supplier_id = s.id
        LEFT JOIN deliverables d ON c.id = d.contract_id
        LEFT JOIN payments pm ON d.id = pm.deliverable_id
        LEFT JOIN warranties w ON c.id = w.contract_id
        ORDER BY p.proj_id_no
    """
    
    try:
        cur.execute(query)
        rows = cur.fetchall()
        
        # Apply filters programmatically
        filtered_rows = []
        if filters:
            from datetime import datetime, date
            status_filter = filters.get("status", "All Statuses")
            division_filter = filters.get("division", "All Divisions")
            date_filter = filters.get("date", "All Dates")
            current_date = date.today()
            
            for row in rows:
                row_status = row[7] # index 0 is p.id, so shifted by 1
                row_division = row[3]
                row_date_str = row[24]
                
                # Status Filter
                if status_filter != "All Statuses" and row_status != status_filter:
                    continue
                    
                # Division Filter
                if division_filter != "All Divisions" and (row_division is None or row_division.strip() != division_filter.strip()):
                    continue
                    
                # Date Presets Filter
                if date_filter != "All Dates" and row_date_str:
                    try:
                        proj_date = datetime.strptime(row_date_str, "%Y-%m-%d").date()
                        days_diff = (current_date - proj_date).days
                        if date_filter == "Within 24 hours only" and not (0 <= days_diff <= 1):
                            continue
                        elif date_filter == "Within a week only" and not (0 <= days_diff <= 7):
                            continue
                        elif date_filter == "Within a month only" and not (0 <= days_diff <= 30):
                            continue
                    except Exception:
                        pass
                
                # Append normal columns (row[1:24]) and then the 9 doc fields
                pid = row[0]
                base_data = list(row[1:24])
                for dtype in ["SARO", "PPMP", "RQ", "MS", "TS", "PO_Phase2", "RFQ_Phase3", "Abstract", "PO_Phase6"]:
                    base_data.append(project_docs_map.get(pid, {}).get(dtype, ""))
                filtered_rows.append(base_data)
        else:
            for row in rows:
                pid = row[0]
                base_data = list(row[1:24])
                for dtype in ["SARO", "PPMP", "RQ", "MS", "TS", "PO_Phase2", "RFQ_Phase3", "Abstract", "PO_Phase6"]:
                    base_data.append(project_docs_map.get(pid, {}).get(dtype, ""))
                filtered_rows.append(base_data)
                
        headers = [
            "Project ID", "Project Name", "Bureau/Division", "Focal Person", "Mode of Procurement", "ABC Budget (₱)", "Project Status",
            "Budget Type", "App Allocation (₱)", "ORS Amount (₱)", "ORS Serial No",
            "Supplier Name", "PO/JO Contract No", "Contract Amount (₱)",
            "Milestone Deliverable", "Delivery Status", "Actual Delivery Date",
            "Payment Type", "Payment Gross (₱)", "Payment Net (₱)", "DV / RCI No", "Check Date",
            "Warranty Status",
            "SARO Link", "PPMP Link", "RQ Link", "MS Link", "TS Link", "PO Phase 2 Link", "RFQ Phase 3 Link", "Abstract Link", "PO Phase 6 Link"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Procurement Registry"
        ws.views.sheetView[0].showGridLines = True
        
        create_formatted_sheet(ws, headers, filtered_rows)
        wb.save(output_path)
        
        return True, len(filtered_rows)
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()

def export_project_timeline(project_id, output_path):
    """
    Queries all timeline data related to a specific project and exports it to a styled native Excel worksheet (.xlsx).
    """
    db_file = Path(__file__).parent / "procurement.db"
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    
    # Query project documents map
    try:
        cur.execute("SELECT project_id, document_type, file_reference FROM project_documents WHERE project_id = ?", (project_id,))
        project_docs_map = {}
        for pid, dtype, ref in cur.fetchall():
            if pid not in project_docs_map:
                project_docs_map[pid] = {}
            project_docs_map[pid][dtype] = ref
    except Exception:
        project_docs_map = {}
        
    query = """
        SELECT 
            p.id,
            p.proj_id_no, p.project_name, p.bureau_division_name, p.focal_person, p.mode_of_procurement, p.abc_amount, p.status,
            pb.budget_type, pb.app_amount, pb.ors_amount, pb.ors_serial_no,
            s.supplier_name,
            c.po_jo_contract_no, c.contract_amount,
            d.milestone_deliverable, d.status_of_delivery, d.actual_delivery_date,
            pm.types_of_payment, pm.payment_amount_gross, pm.payment_amount_net, pm.rci_dv_acic_no, pm.check_date_lddap_ada,
            w.retention_period_warranty_status
        FROM projects p
        LEFT JOIN project_budgets pb ON p.id = pb.project_id
        LEFT JOIN contracts c ON p.id = c.project_id
        LEFT JOIN suppliers s ON c.supplier_id = s.id
        LEFT JOIN deliverables d ON c.id = d.contract_id
        LEFT JOIN payments pm ON d.id = pm.deliverable_id
        LEFT JOIN warranties w ON c.id = w.contract_id
        WHERE p.id = ?
        ORDER BY p.proj_id_no
    """
    
    try:
        cur.execute(query, (project_id,))
        rows = cur.fetchall()
        
        filtered_rows = []
        for row in rows:
            pid = row[0]
            base_data = list(row[1:])
            for dtype in ["SARO", "PPMP", "RQ", "MS", "TS", "PO_Phase2", "RFQ_Phase3", "Abstract", "PO_Phase6"]:
                base_data.append(project_docs_map.get(pid, {}).get(dtype, ""))
            filtered_rows.append(base_data)
            
        headers = [
            "Project ID", "Project Name", "Bureau/Division", "Focal Person", "Mode of Procurement", "ABC Budget (₱)", "Project Status",
            "Budget Type", "App Allocation (₱)", "ORS Amount (₱)", "ORS Serial No",
            "Supplier Name", "PO/JO Contract No", "Contract Amount (₱)",
            "Milestone Deliverable", "Delivery Status", "Actual Delivery Date",
            "Payment Type", "Payment Gross (₱)", "Payment Net (₱)", "DV / RCI No", "Check Date",
            "Warranty Status",
            "SARO Link", "PPMP Link", "RQ Link", "MS Link", "TS Link", "PO Phase 2 Link", "RFQ Phase 3 Link", "Abstract Link", "PO Phase 6 Link"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Timeline"
        ws.views.sheetView[0].showGridLines = True
        
        create_formatted_sheet(ws, headers, filtered_rows)
        wb.save(output_path)
        
        return True, len(filtered_rows)
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()
